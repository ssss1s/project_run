from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
from item.models import CollectibleItem
from rest_framework import status, viewsets


from item.serializers import CollectibleItemSerializer


@api_view(['POST'])
@parser_classes([MultiPartParser])
def upload_file(request):
    if 'file' not in request.FILES:
        return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    if not file.name.endswith('.xlsx'):
        return Response({"error": "Only XLSX files are supported"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = load_workbook(filename=file)
        ws = wb.active
        invalid_rows = []

        # 1. Заранее загружаем ВСЕ существующие UID из БД (в нижнем регистре)
        existing_uids = {uid.lower() for uid in CollectibleItem.objects.values_list('uid', flat=True)}
        valid_names = {'Coin', 'Flag', 'Sun', 'Key', 'Bottle', 'Horn'}
        seen_uids_in_file = set()

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1 or not any(row):
                continue

            row_data = list(row)
            try:
                # 2. Проверка наличия всех 6 полей
                if len(row_data) < 6:
                    raise ValueError("Требуется ровно 6 полей")

                # 3. Очистка и подготовка данных
                name = str(row_data[0]).strip()
                uid = str(row_data[1]).strip().lower()  # Нормализуем UID

                # 4. СТРОГАЯ проверка имени (регистрозависимая)
                if name not in valid_names:
                    raise ValueError(f"Недопустимое имя: {name}. Допустимые: {', '.join(valid_names)}")

                # 5. ТОЧНАЯ проверка формата UID (8 hex-символов)
                if len(uid) != 8 or not all(c in '0123456789abcdef' for c in uid):
                    raise ValueError("UID должен быть ровно 8 символов (0-9, a-f)")

                # 6. Проверка на дубликаты В ФАЙЛЕ
                if uid in seen_uids_in_file:
                    raise ValueError(f"UID {uid} дублируется в этом файле")

                # 7. Проверка на существование В БАЗЕ ДАННЫХ
                if uid in existing_uids:
                    raise ValueError(f"UID {uid} уже существует в системе")

                seen_uids_in_file.add(uid)

                # 8. Проверка value (целое положительное число)
                try:
                    value = int(float(str(row_data[2])))  # Двойное преобразование для безопасности
                    if value <= 0:
                        raise ValueError("Значение должно быть положительным")
                except (ValueError, TypeError):
                    raise ValueError("Некорректное числовое значение")

                # 9. Проверка координат (точные Decimal значения)
                try:
                    lat = Decimal(str(row_data[3]).replace(',', '.'))
                    lon = Decimal(str(row_data[4]).replace(',', '.'))
                    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
                        raise ValueError("Координаты вне допустимого диапазона")
                except (InvalidOperation, ValueError):
                    raise ValueError("Некорректный формат координат")

                # 10. СТРОГАЯ проверка URL
                picture = str(row_data[5]).strip().rstrip(';')
                if not (picture.startswith(('http://', 'https://')) and ' ' not in picture):
                    raise ValueError("URL должен начинаться с http:// или https://")

                # Если ВСЕ проверки пройдены - сохраняем
                CollectibleItem.objects.create(
                    name=name,
                    uid=uid,
                    value=value,
                    latitude=lat,
                    longitude=lon,
                    picture=picture
                )

            except Exception as e:
                # Добавляем ТОЛЬКО если есть ошибка валидации
                invalid_rows.append(row_data)

        return Response(invalid_rows, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer
